# for handling spreedsheet we use openpyxl  module
import openpyxl

#load file
path = r"C:\Users\Mehme\PycharmProjects\pythonProject2\Sales_Data.xlsx"
saleData = openpyxl.load_workbook(path)
# use dictionary to store all the date
products = saleData["Sheet1"]
print(products.max_row)     # we have 29 rows on above file, need to iterate each cell for each row:
products_of_each_vender = {}   # dictionary
for product_each_row in range(2,products.max_row+1):
    vendor = products.cell(product_each_row,1).value
    productNo = products.cell(product_each_row,2).value
    quantity = products.cell(product_each_row, 3).value
    price= products.cell(product_each_row, 4).value
    total = products.cell(product_each_row, 5).value
    products_of_each_vender = {"Vendro_Name":vendor, "Product_No":productNo, "Quantity":quantity, "Price":price, "Total":total}


print(products_of_each_vender)